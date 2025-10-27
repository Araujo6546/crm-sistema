from flask import Blueprint, request, jsonify, send_file
from werkzeug.utils import secure_filename
import os
from models.cliente import Cliente
from models.user import db
import tempfile
from openpyxl import Workbook, load_workbook

upload_bp = Blueprint('upload', __name__)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@upload_bp.route('/upload-clientes', methods=['POST'])
def upload_clientes():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': 'Tipo de arquivo não permitido'}), 400
        
        # Obter modo de upload (add ou replace)
        upload_mode = request.form.get('mode', 'add')  # Default: adicionar
        
        # Salvar arquivo temporariamente
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        
        try:
            # Ler planilha Excel usando openpyxl
            wb = load_workbook(temp_path)
            
            # Tentar encontrar a aba de clientes
            sheet_names = ['Clientes', 'CLIENTES', 'clientes', 'Sheet1', 'Planilha1']
            ws = None
            for sheet_name in sheet_names:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    break
            
            if ws is None:
                return jsonify({
                    'success': False, 
                    'message': 'Aba "Clientes" não encontrada na planilha'
                }), 400
            
            # Ler cabeçalhos (primeira linha)
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().upper())
            
            # Mapear colunas esperadas
            column_mapping = {
                'NOME': ['NOME', 'NOME_CLIENTE', 'CLIENTE', 'RAZAO_SOCIAL'],
                'COD_CLIENTE': ['COD_CLIENTE', 'CODIGO', 'CODIGO_CLIENTE', 'ID'],
                'MUNICIPIO': ['MUNICIPIO', 'CIDADE'],
                'FILIAL': ['FILIAL'],
                'CLASSE': ['CLASSE'],
                'POTENCIAL_PECAS': ['POTENCIAL_PECAS', 'POTENCIAL PECAS', 'PECAS'],
                'POTENCIAL_SERVICO': ['POTENCIAL_SERVICO', 'POTENCIAL SERVICO', 'SERVICO'],
                'STATUS_6M': ['STATUS_6M', 'STATUS 6M', 'STATUS'],
                'CONSULTOR_PECAS': ['CONSULTOR_PECAS', 'CONSULTOR PECAS', 'VENDEDOR_PECAS'],
                'CONSULTOR_SERVICOS': ['CONSULTOR_SERVICOS', 'CONSULTOR SERVICOS', 'VENDEDOR_SERVICOS']
            }
            
            # Encontrar índices das colunas
            column_indices = {}
            for field, possible_names in column_mapping.items():
                for i, header in enumerate(headers):
                    if header in possible_names:
                        column_indices[field] = i
                        break
            
            # Verificar se encontrou as colunas essenciais
            required_fields = ['NOME', 'COD_CLIENTE']
            missing_fields = [field for field in required_fields if field not in column_indices]
            if missing_fields:
                return jsonify({
                    'success': False,
                    'message': f'Colunas obrigatórias não encontradas: {", ".join(missing_fields)}'
                }), 400
            
            # Limpar dados existentes apenas se modo for 'replace'
            if upload_mode == 'replace':
                Cliente.query.delete()
                db.session.commit()
            
            # Processar dados
            imported = 0
            updated = 0
            errors = 0
            skipped = 0  # Para modo 'add' quando cliente já existe
            
            for row_num in range(2, ws.max_row + 1):  # Começar da linha 2 (pular cabeçalho)
                try:
                    row = [cell.value for cell in ws[row_num]]
                    
                    # Extrair dados da linha
                    nome = str(row[column_indices['NOME']]) if column_indices.get('NOME') is not None and row[column_indices['NOME']] else None
                    cod_cliente = row[column_indices['COD_CLIENTE']] if column_indices.get('COD_CLIENTE') is not None else None
                    
                    if not nome or not cod_cliente:
                        continue
                    
                    # Converter código do cliente para inteiro
                    try:
                        cod_cliente = int(float(cod_cliente))
                    except (ValueError, TypeError):
                        errors += 1
                        continue
                    
                    # Extrair outros campos opcionais
                    municipio = str(row[column_indices['MUNICIPIO']]) if column_indices.get('MUNICIPIO') is not None and row[column_indices['MUNICIPIO']] else None
                    filial = str(row[column_indices['FILIAL']]) if column_indices.get('FILIAL') is not None and row[column_indices['FILIAL']] else None
                    classe = str(row[column_indices['CLASSE']]) if column_indices.get('CLASSE') is not None and row[column_indices['CLASSE']] else None
                    
                    # Potenciais (converter para float)
                    potencial_pecas = 0.0
                    potencial_servico = 0.0
                    
                    if column_indices.get('POTENCIAL_PECAS') is not None and row[column_indices['POTENCIAL_PECAS']]:
                        try:
                            potencial_pecas = float(row[column_indices['POTENCIAL_PECAS']])
                        except (ValueError, TypeError):
                            pass
                    
                    if column_indices.get('POTENCIAL_SERVICO') is not None and row[column_indices['POTENCIAL_SERVICO']]:
                        try:
                            potencial_servico = float(row[column_indices['POTENCIAL_SERVICO']])
                        except (ValueError, TypeError):
                            pass
                    
                    status_6m = str(row[column_indices['STATUS_6M']]) if column_indices.get('STATUS_6M') is not None and row[column_indices['STATUS_6M']] else None
                    consultor_pecas = str(row[column_indices['CONSULTOR_PECAS']]) if column_indices.get('CONSULTOR_PECAS') is not None and row[column_indices['CONSULTOR_PECAS']] else None
                    consultor_servicos = str(row[column_indices['CONSULTOR_SERVICOS']]) if column_indices.get('CONSULTOR_SERVICOS') is not None and row[column_indices['CONSULTOR_SERVICOS']] else None
                    
                    # Verificar se cliente já existe
                    existing_cliente = Cliente.query.filter_by(cod_cliente=cod_cliente).first()
                    
                    if existing_cliente:
                        if upload_mode == 'add':
                            # No modo 'add', pular clientes existentes
                            skipped += 1
                            continue
                        else:
                            # No modo 'replace', atualizar cliente existente
                            existing_cliente.nome = nome
                            existing_cliente.municipio = municipio
                            existing_cliente.filial = filial
                            existing_cliente.classe = classe
                            existing_cliente.potencial_pecas = potencial_pecas
                            existing_cliente.potencial_servico = potencial_servico
                            existing_cliente.status_6m = status_6m
                            existing_cliente.consultor_pecas = consultor_pecas
                            existing_cliente.consultor_servicos = consultor_servicos
                            updated += 1
                    else:
                        # Criar novo cliente
                        cliente = Cliente(
                            nome=nome,
                            cod_cliente=cod_cliente,
                            municipio=municipio,
                            filial=filial,
                            classe=classe,
                            potencial_pecas=potencial_pecas,
                            potencial_servico=potencial_servico,
                            status_6m=status_6m,
                            consultor_pecas=consultor_pecas,
                            consultor_servicos=consultor_servicos
                        )
                        db.session.add(cliente)
                        imported += 1
                        
                except Exception as e:
                    print(f"Erro ao processar linha {row_num}: {str(e)}")
                    errors += 1
                    continue
            
            # Salvar alterações
            db.session.commit()
            
            # Preparar mensagem baseada no modo
            if upload_mode == 'add':
                message = f'Clientes adicionados com sucesso! {skipped} clientes já existentes foram ignorados.'
            else:
                message = f'Base de clientes substituída com sucesso!'
            
            return jsonify({
                'success': True,
                'message': message,
                'stats': {
                    'imported': imported,
                    'updated': updated,
                    'skipped': skipped if upload_mode == 'add' else 0,
                    'total': imported + updated,
                    'errors': errors,
                    'mode': upload_mode
                }
            })
            
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Erro ao processar arquivo: {str(e)}'
            }), 500
            
        finally:
            # Remover arquivo temporário
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro interno: {str(e)}'
        }), 500

@upload_bp.route('/download-template', methods=['GET'])
def download_template():
    try:
        # Criar workbook com template
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Cabeçalhos
        headers = [
            'NOME', 'COD_CLIENTE', 'MUNICIPIO', 'FILIAL', 'CLASSE',
            'POTENCIAL_PECAS', 'POTENCIAL_SERVICO', 'STATUS_6M',
            'CONSULTOR_PECAS', 'CONSULTOR_SERVICOS'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Exemplo de dados
        example_data = [
            'EMPRESA EXEMPLO LTDA', 123456, 'São Paulo', 'SP', 'A',
            15000.00, 8000.00, 'ATIVO', 'VENDEDOR EXEMPLO', 'VENDEDOR EXEMPLO'
        ]
        
        for col, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Salvar em arquivo temporário
        temp_path = os.path.join(tempfile.gettempdir(), 'template_clientes.xlsx')
        wb.save(temp_path)
        
        return send_file(
            temp_path,
            as_attachment=True,
            download_name='template_clientes.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao gerar template: {str(e)}'
        }), 500

